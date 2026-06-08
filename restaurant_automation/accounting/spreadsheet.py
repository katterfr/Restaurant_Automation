"""
spreadsheet.py — Auto-generate Excel accounting workbooks with openpyxl.
Produces a multi-sheet XLSX with:
  1. Summary Dashboard
  2. Cash Flow Statement
  3. Inventory Snapshot
  4. Transaction Log
  5. Purchase Orders
  6. P&L Statement
"""
from __future__ import annotations
import logging
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

from accounting.cash_flow import get_cash_flow_statement
from accounting.pnl import get_pnl_statement
from accounting.loss_gain import get_loss_gain_summary
from inventory.inventory_db import get_all_items, get_transaction_history
from inventory.purchase_order import get_all_pos

log = logging.getLogger(__name__)

# ── Style constants ──────────────────────────────────────────────────────────
HEADER_FILL  = PatternFill("solid", fgColor="1F3864")
HEADER_FONT  = Font(color="FFFFFF", bold=True, size=11)
SUBHEAD_FILL = PatternFill("solid", fgColor="2E75B6")
SUBHEAD_FONT = Font(color="FFFFFF", bold=True, size=10)
ALT_FILL     = PatternFill("solid", fgColor="D9E1F2")
POS_FONT     = Font(color="375623", bold=True)
NEG_FONT     = Font(color="C00000", bold=True)
BORDER       = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin")
)
MONEY_FMT    = '$#,##0.00'
PCT_FMT      = '0.00%'
DATE_FMT     = 'YYYY-MM-DD'


def _header(ws, row: int, col: int, value: str, wide: bool = False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill  = HEADER_FILL
    cell.font  = HEADER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER
    return cell


def _subheader(ws, row: int, col: int, value: str):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill  = SUBHEAD_FILL
    cell.font  = SUBHEAD_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = BORDER
    return cell


def _money(ws, row: int, col: int, value: float):
    cell = ws.cell(row=row, column=col, value=value)
    cell.number_format = MONEY_FMT
    cell.alignment = Alignment(horizontal="right")
    cell.border = BORDER
    if value < 0:
        cell.font = NEG_FONT
    elif value > 0:
        cell.font = POS_FONT
    return cell


def _text(ws, row: int, col: int, value: Any):
    cell = ws.cell(row=row, column=col, value=value)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = BORDER
    return cell


def _auto_col_widths(ws, min_width: int = 10, max_width: int = 40):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 2, max_width))


# ── Sheet builders ──────────────────────────────────────────────────────────

def _build_summary(ws, pnl: dict, cf: dict, restaurant_name: str, report_date: str):
    ws.title = "📊 Summary"
    ws.sheet_view.showGridLines = False

    # Title block
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = f"{restaurant_name} — Financial Summary"
    title_cell.font = Font(size=16, bold=True, color="1F3864")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:F2")
    sub = ws["A2"]
    sub.value = f"Report Date: {report_date}"
    sub.font = Font(italic=True, color="595959")
    sub.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    metrics = [
        ("Total Revenue",       pnl.get("total_revenue", 0)),
        ("Cost of Goods Sold",  pnl.get("total_cogs", 0)),
        ("Gross Profit",        pnl.get("gross_profit", 0)),
        ("Total Operating Exp", pnl.get("total_operating_expenses", 0)),
        ("Net Income",          pnl.get("net_income", 0)),
        ("Net Cash Flow",       cf.get("net_change_in_cash", 0)),
    ]

    for i, (label, value) in enumerate(metrics, start=4):
        ws.row_dimensions[i].height = 22
        _subheader(ws, i, 1, label)
        _money(ws, i, 2, value)

    # Simple bar chart placeholder
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 16


def _build_pnl(ws, pnl: dict):
    ws.title = "💰 P&L"
    ws.sheet_view.showGridLines = False
    row = 1

    headers = ["Account Code", "Account Name", "Amount"]
    for c, h in enumerate(headers, 1):
        _header(ws, row, c, h)
    row += 1

    sections = [
        ("Revenue", pnl.get("revenue_lines", [])),
        ("Cost of Goods Sold", pnl.get("cogs_lines", [])),
        ("Operating Expenses", pnl.get("expense_lines", [])),
    ]

    for section_name, lines in sections:
        ws.merge_cells(f"A{row}:C{row}")
        ws.cell(row=row, column=1, value=section_name).font = Font(bold=True, size=11)
        row += 1
        total = 0.0
        for line in lines:
            _text(ws, row, 1, line.get("code", ""))
            _text(ws, row, 2, line.get("name", ""))
            _money(ws, row, 3, line.get("amount", 0))
            total += line.get("amount", 0)
            row += 1
        # Section total
        ws.cell(row=row, column=2, value=f"Total {section_name}").font = Font(bold=True)
        _money(ws, row, 3, total)
        row += 2

    # Net Income
    ws.cell(row=row, column=2, value="NET INCOME").font = Font(bold=True, size=12)
    _money(ws, row, 3, pnl.get("net_income", 0))
    _auto_col_widths(ws)


def _build_cash_flow(ws, cf: dict):
    ws.title = "💵 Cash Flow"
    ws.sheet_view.showGridLines = False
    row = 1

    _header(ws, row, 1, "Section")
    _header(ws, row, 2, "Account")
    _header(ws, row, 3, "Description")
    _header(ws, row, 4, "Amount")
    row += 1

    op = cf.get("operating", {})
    for label, items in [
        ("Operating Inflows",  op.get("inflows", [])),
        ("Operating Outflows", op.get("outflows", [])),
    ]:
        for item in items:
            _text(ws, row, 1, label)
            _text(ws, row, 2, item.get("account"))
            _text(ws, row, 3, item.get("name"))
            _money(ws, row, 4, item.get("amount", 0))
            row += 1

    row += 1
    ws.cell(row=row, column=3, value="Net Operating Cash Flow").font = Font(bold=True)
    _money(ws, row, 4, op.get("net_operating", 0))
    row += 2
    ws.cell(row=row, column=3, value="NET CHANGE IN CASH").font = Font(bold=True, size=12)
    _money(ws, row, 4, cf.get("net_change_in_cash", 0))
    _auto_col_widths(ws)


def _build_inventory(ws, items: list[dict]):
    ws.title = "📦 Inventory"
    ws.sheet_view.showGridLines = False
    row = 1

    headers = ["SKU", "Name", "Category", "Unit", "Qty On Hand",
               "Reorder Level", "Cost/Unit", "Sell Price",
               "Inventory Value", "Status"]
    for c, h in enumerate(headers, 1):
        _header(ws, row, c, h)
    row += 1

    for i, item in enumerate(items, 1):
        fill = ALT_FILL if i % 2 == 0 else None
        inv_value = round(item["qty_on_hand"] * item["cost_per_unit"], 2)
        status = "⚠️ LOW" if item["qty_on_hand"] <= item["reorder_level"] else "✅ OK"

        for c, val in enumerate([
            item["sku"], item["name"], item["category"], item["unit"],
            item["qty_on_hand"], item["reorder_level"],
            item["cost_per_unit"], item["sell_price"], inv_value, status
        ], 1):
            cell = ws.cell(row=row, column=c, value=val)
            if fill:
                cell.fill = fill
            cell.border = BORDER
            if c in (5, 6):
                cell.number_format = "#,##0.00"
            elif c in (7, 8, 9):
                cell.number_format = MONEY_FMT
        row += 1

    _auto_col_widths(ws)


def _build_transactions(ws, txns: list[dict]):
    ws.title = "📋 Transactions"
    ws.sheet_view.showGridLines = False
    row = 1

    headers = ["Date", "SKU", "Item Name", "Delta", "Reason", "Order Ref", "Note"]
    for c, h in enumerate(headers, 1):
        _header(ws, row, c, h)
    row += 1

    for txn in txns[:500]:   # cap at 500 rows for performance
        vals = [
            txn.get("created_at", "")[:10],
            txn.get("sku", ""),
            txn.get("name", ""),
            txn.get("delta", 0),
            txn.get("reason", ""),
            txn.get("order_ref", ""),
            txn.get("note", ""),
        ]
        for c, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.border = BORDER
            if c == 4:
                cell.number_format = "+#,##0.00;-#,##0.00"
                cell.font = POS_FONT if float(val or 0) >= 0 else NEG_FONT
        row += 1

    _auto_col_widths(ws)


def _build_purchase_orders(ws, pos: list[dict]):
    ws.title = "🛒 Purchase Orders"
    ws.sheet_view.showGridLines = False
    row = 1

    headers = ["PO Number", "Supplier", "Status", "Total Cost",
               "Ordered At", "Received At"]
    for c, h in enumerate(headers, 1):
        _header(ws, row, c, h)
    row += 1

    for po in pos:
        vals = [
            po.get("po_number"), po.get("supplier"), po.get("status"),
            po.get("total_cost", 0),
            po.get("ordered_at", "")[:10],
            po.get("received_at", "") or "",
        ]
        for c, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.border = BORDER
            if c == 4:
                cell.number_format = MONEY_FMT
        row += 1

    _auto_col_widths(ws)


# ── Main entry point ─────────────────────────────────────────────────────────

async def generate_workbook(
    start_date: date,
    end_date: date,
    restaurant_name: str,
    output_path: str | Path,
    db_path: str = "restaurant.db",
) -> Path:
    """
    Generate the full accounting workbook and save to output_path.
    Returns the Path to the written file.
    """
    log.info("Generating workbook for %s → %s", start_date, end_date)

    # Fetch all data
    pnl     = await get_pnl_statement(start_date, end_date, db_path=db_path)
    cf      = await get_cash_flow_statement(start_date, end_date, db_path=db_path)
    items   = await get_all_items(db_path=db_path)
    txns    = await get_transaction_history(limit=500, db_path=db_path)
    pos     = await get_all_pos(db_path=db_path)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default sheet

    report_date = datetime.now(timezone.utc).date().isoformat()

    _build_summary(wb.create_sheet(), pnl, cf, restaurant_name, report_date)
    _build_pnl(wb.create_sheet(), pnl)
    _build_cash_flow(wb.create_sheet(), cf)
    _build_inventory(wb.create_sheet(), items)
    _build_transactions(wb.create_sheet(), txns)
    _build_purchase_orders(wb.create_sheet(), pos)

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    log.info("Workbook saved → %s", out)
    return out
